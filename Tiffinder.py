""" Find *.tif[f] files by identNr 

    Uses a json file as cache to store path information (e.g. .tif_cache.json)

    For command-line front-end see find_tif.py.

    USAGE as class:
        tf = Tiffinder(cache_fn)

        #working with the cache
        tf.scandir(scan_dir)  # scans recursively for *.tif|*.tiff
        tf.iscandir(scan_dir) # rescans dir if cache is older than 1d
        tf.show_cache()       # prints cache to STDOUT 

        #three different searches
        r = tf.search (needle) # matching path in list, for single needle
        r = tf.search_xlsx(needle) # with multiple needles from xls, 
                              # search cache & report to STDOUT

        r = tf.search_xlsx(xls_fn, outdir) 
                              # search cache for needles from xls, copy found 
                              # tifs to outdir
                              # output: orig-filename (no).tif
        r = tf.search_mpx(mpx_fn, outdir) 
                              # search for all identNr in mpx, copy to outdir
                              # output: objId.hash.tif

        #do stuff with search results
        tf.cp_results(results, target_dir, [policy])
        tf.log_results(results, [policy])
        tf.preview_results(results, target_dir, [policy]) 

    Filename Policy
    - default: Preserve except if not unique
    - hash: objId.hash.tif (not implemented anymore)
    - DAid: DAid.origname.tif (not implemented yet)

    When a file is copied, the original filename is usually preserved; only if 
    multiple tifs have the same name they are varied by adding a number. The 
    downside of this naming scheme is that if tiffinder runs multiple times, 
    same files will be copied multiple times (because there is no identity). This 
    naming scheme may be useful for some uses, just beware.
    
    search_mpx uses other naming scheme:
        objId.hash.tif
        
    MulId naming scheme
        DAid.origname.tif
"""

import datetime
import filecmp
import json
import logging
import shutil
import time
from lxml import etree
from PIL import Image
from openpyxl import Workbook, load_workbook
from pathlib import Path


class Tiffinder:
    def __init__(self, cache_fn):
        """initialize object

        cache_fn: location (path) of cache file"""

        self.cache_fn = Path(cache_fn)
        # print ('cache_fn %s' % cache_fn)

        if Path(self.cache_fn).exists():
            print(f"*Cache exists, loading '{self.cache_fn}'")
            with open(self.cache_fn, "r") as f:
                self.cache = json.load(f)
        else:
            self.cache = {}

    def scandir(self, scan_dir):
        """Scan dir for tif and store result in tif cache.

        Scans directory for *.tif[f?] recursively and write results to
        cache file, updating existing cache file or starting new one.

        Does a sloppy update, i.e will not remove cache entries for files that
        have been removed from disk. See iscandir to avoid that.

        Repeat to scan multiple dirs."""

        print(f"*Scanning '{scan_dir}'")
        for path in Path(scan_dir).rglob("**/*.tif[f?]"): 
            abs = Path(path).resolve()
            needle = abs.stem.replace("_", " ")
            # VII c 123 a,b <1>
            # only 5 elements, discard rest
            short = " ".join(needle.split(" ", maxsplit=5)) 
            print(f"{abs} - {short}")
            self.cache[str(abs)] = short
        self._write_cache()

    def iscandir(self, scan_dir):
        """intelligent directory scan

        scan_dir can be a list

        Do the same scan as in scandir, but also remove cache items whose files
        don't exist on disk anymore and do a repeated scan only if cache is
        older than 10 days.

        Scan multiple directories by passing a list:
            tf.iscandir (['.', '.'])
        """

        print(f"*About to i-scan {scan_dir}")
        cache_mtime = self.cache_fn.lstat().st_mtime
        now = time.time()
        # print (f"DIFF:{now-cache_mtime}")
        # only if cache is older than one day
        if now - cache_mtime > 10 * 3600 * 24:
            # remove items from cache if file doesn't exist anymore
            for path in list(self.cache):
                if not os.path.exists(path):
                    print(f"File doesn't exist anymore; remove from cache {path}")
                    del self.cache[path]
            for dir in scan_dir:
                self.scandir(dir)  # writes cache
        else:
            print("*Cache still young")

    def search(self, needle):
        """Search tif cache for a single needle.

        Return list of matches:
            ls=self.search(needle)"""

        # print ("* Searching cache for needle '%s'" % needle)
        return [path for path in self.cache if needle in self.cache[path]]

    def search_xlsx(self, xls_fn):
        """Search tif cache for needles from Excel file.

        Needles are coming from xlsx filed (first sheet, column A)"""

        print(f"* Searching cache for needles from Excel file {xls_fn}")

        self.wb = self._prepare_wb(xls_fn)
        # ws = self.wb.active # last active sheet
        ws = self.wb.worksheets[0]
        print(f"* Sheet title: {ws.title}")
        col = ws["A"]  # zero or one based?
        results = []
        for needle in col:
            if needle.value is not None:
                print(f"* Looking for '{needle.value}'")
                found = self.search(needle.value)
                print(f"* FOUND: {len(found)}")
                for each in found:
                    print(f"\teach")
                results.extend(found)
        return results

    def search_mpx(self, mpx_fn):
        """Search tif cache for identNr from mpx.

        For each identNr from mpx look for corresponding tifs in cache"""

        tree = etree.parse(mpx_fn)
        r = tree.xpath(
            "/m:museumPlusExport/m:sammlungsobjekt/m:identNr",
            namespaces={"m": "http://www.mpx.org/mpx"},
        )

        results = []
        for identNr_node in r:
            tifs = self.search(identNr_node.text)
            objId = tree.xpath(
                f"/m:museumPlusExport/m:sammlungsobjekt/@objId[../m:identNr = '{identNr_node.text}']",
                namespaces={"m": "http://www.mpx.org/mpx"},
            )[0]
            # print(f"{identNr_node.text}->{objId}")
            found = self.search(identNr_node.text)
            for f in found:
                print(f"{identNr_node.text}->{objId}->{f}")
                results.extend(found)
            return results

    def show_cache(self):
        """Prints contents of cache to STDOUT"""

        print("*Displaying cache contents")
        if hasattr(self, "cache"):
            for item in self.cache:
                print(f" {item} [{self.cache[item]}]")
            print(f"Number of tifs in cache: {len(self.cache)}")
        else:
            print("Cache does not exist!")

    #
    # output
    #
    def cp_results(self, results, target_dir, policy=None):
        """ Copy files from the search_results to target_dir."""

        self._init_log(target_dir)
        for fn in results:
            target_fn = self._default_policy(fn, target_dir)
            print(f"{fn}->")
            if target_fn is None:
                print(" identical file already exists at target")
            else:
                print(f" {target_fn}")
                self._simple_copy(fn, target_fn)

    def log_results(self, results, target_dir):
        """ Just print the log to target_dir, don't copy anything."""

        self._init_log(target_dir)
        for fn in results:
            print(fn)
            if not Path(fn).exists():
                logging.debug(f"File not found: {fn}")

    def preview_results(self, results, target_dir):
        """ Make previews (720 px wide jpgs) for the search results and copy 
        those to target_dir.

        Always uses change extension naming policy.
        """

        self._init_log(target_dir)
        for fn in results:
            target_fn = self._change_extension_policy(fn, target_dir, "jpg")
            print(f"***{target_fn}")
            self._preview(fn, target_fn)

    ############# PRIVATE STUFF #############

    def _default_policy(self, source, target_dir):
        """Returns target filename as pathlib object or None if identical
        file is already present at target_dir.

        If non identical files with the same exists already in target_dir,
        a numbered variant is returned.
            target_dir/name.suffix
            target_dir/name (1).suffix
            ...
        Expects a full path as str for source, returns Pathlib object.
        """

        i = 2
        new = Path(source)
        suffix = Path(new).suffix
        parent = new.parent
        target_fn = Path(target_dir).joinpath(new.name)
        # print(f"TTT:{target_fn}")
        while target_fn.exists():
            if filecmp.cmp(source, target_fn, shallow=False):
                print(f"identical file exists already in target_dir, no copy")
                return None
            else:
                target_fn = parent.joinpath(f"{new.stem} ({i}){suffix}")
            i += 1
        # print(f"POLICY: [{i}] {new}")
        return target_fn

    def _change_extension_policy(self, source, target_dir, new_ext):
        """Like default policy just without identity check.

        Note: if you run this command multiple times you'll get duplicates
        """

        i = 2
        source = Path(source)
        target_dir = Path(target_dir)
        suffix = "." + new_ext
        target_fn = target_dir.joinpath(source.stem + suffix)
        while target_fn.exists():  # why while?
            target_fn = target_dir.joinpath(f"{source.stem} ({i}){suffix}")
            i += 1
        return target_fn

    def _init_log(self, outdir="."):
        log_fn = Path(outdir).joinpath("tiffinder.log")

        logging.basicConfig(
            datefmt="%Y%m%d %I:%M:%S %p",
            filename=log_fn,
            filemode="w",
            level=logging.DEBUG,
            format="%(asctime)s: %(message)s",
        )

    def _simple_copy(self, source, target):
        """Copy source to target. Expects full paths."""

        # print(f"{source} ->\n\t{target}")
        if not target.is_file():  # no overwrite
            try:
                shutil.copy2(source, target)  # copy2 preserves file info
            except:
                logging.debug(f"File not found: {source}")

    def _prepare_wb(self, xls_fn):
        """Read existing xlsx and return workbook"""

        if Path(xls_fn).is_file():
            # print (f"File exists ({xls_fn})")
            return load_workbook(filename=xls_fn)
        else:
            raise ValueError(f"Excel file not found: {xls_fn}")

    def _write_cache(self):
        print("*Writing cache file")
        with open(self.cache_fn, "w") as f:
            json.dump(self.cache, f, indent=1)

    def _preview(self, source, target):
        im = Image.open(source)

        if im.height > 720:
            ratio = 720 / im.height
            new_size = round(im.width * ratio), round(im.height * ratio)
            print(f"{source}: ({im.width}, {im.height}) -> {new_size} {ratio}")
            im.thumbnail(new_size)
        rgb_im = im.convert("RGB")
        print(f" saving {target}")
        rgb_im.save(str(target))

""" Find *.tif[f] files by identNr 

    Uses a json file as cache to store path information (e.g. .tif_cache.json)

    For command-line front-end see end of this file.

    USAGE as class:
        tf = Tif_finder(cache_fn)

        #working with the cache
        tf.scandir(scan_dir)  # scans recursively for *.tif|*.tiff
        tf.iscandir(scan_dir) # rescans dir if cache is older than 1d
        tf.show_cache()       # prints cache to STDOUT 

        #three different searches
        ls=tf.search (needle) # matching path in list, for single needle
        r = tf.search_xlsx(needle) # with multiple needles from xls, 
                              # search cache & report to STDOUT

        r = tf.search_xlsx(xls_fn, outdir) 
                              # search cache for needles from xls, copy found 
                              # tifs to outdir
                              # output: orig-filename (no).tif
        r = tf.search_mpx(mpx_fn, outdir) 
                              # search for all identNr in mpx, copy to outdir
                              # output: objId.hash.tif
        #do stuff with search results
        tf.cp_results(results, target_dir, [policy])
        tf.log_results(results, [policy])
        tf.preview_results(results, target_dir, [policy]) 

    Filename Policy
    - default: Preserve except if not unique
    - hash: objId.hash.tif
    - DAid: DAid.origname.tif

    When a file is copied, the original filename is usually preserved; only if 
    multiple tifs have the same name they are varied by adding a number. The 
    downside of this naming scheme is that if Tif_finder runs multiple times, 
    same files will be copied multiple times (because there is no identity). This 
    naming scheme may be useful for some uses, just beware.
    
    search_mpx uses other naming scheme:
        objId.hash.tif
        
    MulId naming scheme
        DAid.origname.tif
"""

import datetime
import filecmp
import hashlib
import json
import logging
import pprint
from lxml import etree

# import os
import shutil
import time
from PIL import Image
from openpyxl import Workbook, load_workbook
from pathlib import Path
from glob import iglob


class Tif_finder:
    def __init__(self, cache_fn):
        """initialize object

        cache_fn: location (path) of cache file"""
        self.cache_fn = Path(cache_fn)
        # print ('cache_fn %s' % cache_fn)

        if Path(self.cache_fn).exists():
            print(f"*Cache exists, loading '{self.cache_fn}'")
            with open(self.cache_fn, "r") as f:
                self.cache = json.load(f)
        else:
            self.cache = {}

    def scandir(self, scan_dir):
        """Scan dir for tif and store result in tif cache.

        Scans directory for *.tif|*.tiff recursively and write results to
        cache file, updating existing cache file or starting new one.

        Does a sloppy update, i.e will not remove cache entries for files that
        have been removed from disk. See iscandir to avoid that.

        Repeat to scan multiple dirs."""
        print(f"*Scanning '{scan_dir}'")
        for path in Path(scan_dir).rglob("**/*.ti[f*]"):
            abs = Path(path).resolve()
            needle = abs.stem.replace("_", " ")
            print(f"{abs}")
            self.cache[str(abs)] = needle
        self._write_cache()

    def iscandir(self, scan_dir):
        """intelligent directory scan

        scan_dir can be a list

        Do the same scan as in scandir, but also remove cache items whose files
        don't exist on disk anymore and do a repeated scan only if cache is
        older than 1 day.

        Scan multiple directories by passing a list:
            tf.iscandir (['.', '.'])
        """

        print(f"*About to i-scan {scan_dir}")
        cache_mtime = self.cache_fn.lstat().st_mtime
        now = time.time()
        # print (f"DIFF:{now-cache_mtime}")
        # only if cache is older than one day
        if now - cache_mtime > 10 * 3600 * 24:
            # remove items from cache if file doesn't exist anymore
            for path in list(self.cache):
                if not os.path.exists(path):
                    print(f"File doesn't exist anymore; remove from cache {path}")
                    del self.cache[path]
            for dir in scan_dir:
                self.scandir(dir)  # writes cache
        else:
            print("*Cache still young")

    def search(self, needle, target_dir=None):
        """Search tif cache for a single needle.

        Return list of matches:
            ls=self.search(needle)

        If target_dir is provided copy matches to that dir."""

        # print ("* Searching cache for needle '%s'" % needle)
        return [path for path in self.cache if needle in self.cache[path]]

    def search_xlsx(self, xls_fn, target_dir=None):
        """Search tif cache for needles from Excel file.

        Needles are expected (first sheet, column A)
        If target_dir is not None, copy the file to respective directory.
        If target_dir is None just report matching paths to STDOUT"""

        print(f"* Searching cache for needles from Excel file {xls_fn}")

        self.wb = self._prepare_wb(xls_fn)
        # ws = self.wb.active # last active sheet
        ws = self.wb.worksheets[0]
        print(f"* Sheet title: {ws.title}")
        col = ws["A"]  # zero or one based?
        results = []
        for needle in col:
            if needle.value is not None:
                print(f"* Looking for '{needle.value}'")
                found = self.search(needle.value)
                print(f"* FOUND: {len(found)}")
                for each in found:
                    print(f"\teach")
                results.extend(found)
        return results

    def search_mpx(self, mpx_fn, target_dir=None):
        """Search tif cache for identNr from mpx.

        For each identNr from mpx look for corresponding tifs in cache; if
        target_dir exists, copy them to target_dir. If target_dir doesn't exist
        just report them to STDOUT."""

        if target_dir is not None:
            target_dir = os.path.realpath(target_dir)
            if not os.path.isdir(target_dir):
                os.makedirs(target_dir)
        tree = etree.parse(mpx_fn)
        r = tree.xpath(
            "/m:museumPlusExport/m:sammlungsobjekt/m:identNr",
            namespaces={"m": "http://www.mpx.org/mpx"},
        )

        results = []
        for identNr_node in r:
            tifs = self.search(identNr_node.text)
            objId = tree.xpath(
                f"/m:museumPlusExport/m:sammlungsobjekt/@objId[../m:identNr = '{identNr_node.text}']",
                namespaces={"m": "http://www.mpx.org/mpx"},
            )[0]
            # print(f"{identNr_node.text}->{objId}")
            found = self.search(identNr_node.text)
            for f in found:
                print(f"{identNr_node.text}->{objId}->{f}")
                results.extend(found)
            return results

    def show_cache(self):
        """Prints contents of cache to STDOUT"""

        print("*Displaying cache contents")
        if hasattr(self, "cache"):
            for item in self.cache:
                print(f"  {item}")
            print(f"Number of tifs in cache: {len(self.cache)}")
        else:
            print("Cache does not exist!")

    #
    # output
    #
    def cp_results(self, results, target_dir, policy=None):
        """ Copy files from the search_results to target_dir."""
        self._init_log(target_dir)

        for fn in results:
            target_fn = self._default_policy(fn, target_dir)
            print(f"{fn}->")
            if target_fn is None:
                print(" identical file already exists at target")
            else:
                print(f" {target_fn}")
                self._simple_copy(fn, target_fn)

    def log_results(self, results, args):
        """ Just print the log to target_dir, don't copy anything."""
        for fn in results:
            print(fn)
            # todo logging

    def preview_results(self, results, target_dir):
        """Make previews for the search results and copy those to target_dir.

        Always uses change extension naming policy.
        """

        for fn in results:
            target_fn = self._change_extension_policy(fn, target_dir, "jpg")
            print(f"***{target_fn}")
            self._preview(fn, target_fn)

    ############# PRIVATE STUFF #############

    def _default_policy(self, source, target_dir):
        """Returns target filename as pathlib object or None if identical
        file is already present at target_dir.

        If non identical files with the same exists already in target_dir,
        a numbered variant is returned.
            target_dir/name.suffix
            target_dir/name (1).suffix
            ...
        Expects a full path as str for source, returns Pathlib object.
        """
        i = 2
        new = Path(source)
        suffix = Path(new).suffix
        parent = new.parent
        target_fn = Path(target_dir).joinpath(new.name)
        # print(f"TTT:{target_fn}")
        while target_fn.exists():
            if filecmp.cmp(source, target_fn, shallow=False):
                print(f"identical file exists already in target_dir, no copy")
                return None
            else:
                target_fn = parent.joinpath(f"{new.stem} ({i}){suffix}")
            i += 1
        # print(f"POLICY: [{i}] {new}")
        return target_fn

    def _change_extension_policy(self, source, target_dir, new_ext):
        """Like default policy just without identity check.

        Note: if you run this command multiple times you'll get duplicates
        """
        i = 2
        source = Path(source)
        target_dir = Path(target_dir)
        suffix = "." + new_ext
        target_fn = target_dir.joinpath(source.stem + suffix)
        while target_fn.exists():  # why while?
            target_fn = target_dir.joinpath(f"{source.stem} ({i}){suffix}")
            i += 1
        return target_fn

    def _init_log(self, outdir):
        log_fn = Path(outdir).joinpath("report.log")

        logging.basicConfig(
            datefmt="%Y%m%d %I:%M:%S %p",
            filename=log_fn,
            filemode="w",
            level=logging.DEBUG,
            format="%(asctime)s: %(message)s",
        )

    def _simple_copy(self, source, target):
        """Copy source to target. Expects full paths."""

        # print(f"{source} ->\n\t{target}")
        if not target.is_file():  # no overwrite
            try:
                shutil.copy2(source, target)  # copy2 preserves file info
            except:
                logging.debug(f"File not found: {source}")

    def _prepare_wb(self, xls_fn):
        """Read existing xlsx and return workbook"""

        if Path(xls_fn).is_file():
            # print (f"File exists ({xls_fn})")
            return load_workbook(filename=xls_fn)
        else:
            raise ValueError(f"Excel file not found: {xls_fn}")

    def _write_cache(self):
        print("*Writing cache file")
        with open(self.cache_fn, "w") as f:
            json.dump(self.cache, f, indent=1)

    def _preview(self, source, target):
        im = Image.open(source)

        if im.height > 720:
            ratio = 720 / im.height
            new_size = round(im.width * ratio), round(im.height * ratio)
            print(f"{source}: ({im.width}, {im.height}) -> {new_size} {ratio}")
            im.thumbnail(new_size)
        rgb_im = im.convert("RGB")
        print(f" saving {target}")
        rgb_im.save(str(target))


if __name__ == "__main__":
    """Find *.tif[f] files by identNr - Command line interface
    USAGE:
    Cache
        #new cache by scanning dir for *.tif[f]
        Tif_finder.py -c cache.json -u scan_dir

        #same, but use intelligent cache update
        Tif_finder.py -c cache.json -i -u scan_dir

        #show cache using specified cache_fn
        Tif_finder.py -c cache.json  -S

    Search
        #Lookup individual needle, cp to pwd
        Tif_finder.py -c cache.json -s needle

        #Lookup individual needle, cp to target_dir
        Tif_finder.py -c cache.json -s needle -t target_dir

        #lookup multiple needles from xlsx file
        Tif_finder.py -c cache.json -x excel_fn

        #read mpx file, lookup all identNr in cache and copy to pwd
        Tif_finder.py -c cache.json -m mpx_fn

    Output
        #normal search, but don't copy anything, just log what would happen
        Tif_finder.py -c cache.json -s neeedle --justlog

        #write preview instead of original file
        Tif_finder.py -c cache.json -s neeedle --preview
    """

    import argparse
    from pathlib import Path

    # from Tif_finder import Tif_finder
    def _output(self, args, r):
        if args.justlog is not False:
            print(f"*JUST LOG")
            self.log_results(r, args.target_dir)
        elif args.preview is not False:
            print(f"*PREVIEW RESULTS")
            self.preview_results(r, args.target_dir)
        else:
            print(f"*COPYING RESULTS")
            self.cp_results(r, args.target_dir)

    parser = argparse.ArgumentParser(
        description="Tif_finder: Find *.tif[f] files by identNr"
    )

    # cache + target_dir
    parser.add_argument("-c", "--cache_fn", required=True, help="Path cache file")
    parser.add_argument(
        "-t",
        "--target_dir",
        default=Path("."),
        help="target directory to write to; defaults to current working directory",
    )
    # main command
    cmd = parser.add_mutually_exclusive_group(required=True)
    cmd.add_argument(
        "-n", "--new_cache", help="Write new cache, expects scan dir", nargs="?"
    )
    cmd.add_argument(
        "-u",
        "--update_cache",
        help="Update existing cache, expects scan dir",
        nargs="?",
    )
    cmd.add_argument(
        "-S", "--show_cache", action="store_true", help="Show cache; no parameter"
    )
    cmd.add_argument(
        "-s", "--search", help="Search for a single needle", nargs="?"
    )  # needle
    cmd.add_argument(
        "-x",
        "--search_xlsx",
        help="Search for all needles in first column of first sheet",
        nargs="?",
    )
    cmd.add_argument(
        "-m", "--search_mpx", help="Search for all needles in mpx/identNr", nargs="?"
    )

    # modify output
    output = parser.add_mutually_exclusive_group()
    output.add_argument(
        "-o", "--output", action="store_true", help="Modify output (preview, log)"
    )
    output.add_argument(
        "-p",
        "--preview",
        action="store_true",
        help="Don't copy tif, but make a smaller preview instead",
    )
    output.add_argument(
        "-j",
        "--justlog",
        action="store_true",
        help="Don't write any image files, just log",
    )

    args = parser.parse_args()
    print(args)
    print(f"*Loading specified cache '{args.cache_fn}'")

    t = Tif_finder(args.cache_fn)

    if args.new_cache is not None:
        t.scandir(args.new_cache)
    elif args.update_cache is not None:
        t.iscandir(args.update_cache)
    elif args.show_cache is not False:
        t.show_cache()
    elif args.search is not None:
        print(f"*Searching for '{args.search}'")
        r = t.search(args.search)
        _output(t, args, r)
    elif args.search_xlsx is not None:
        print("*Excel input")
        r = t.search_xlsx(args.search_xlsx)
        _output(t, args, r)
    elif args.search_mpx is not None:
        print("*Excel input")
        r = t.search_mpx(args.search_mpx)
        _output(t, args, r)
